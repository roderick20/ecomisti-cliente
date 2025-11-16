from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

from modules.facturacion.models.operacion_model import Operacion
from modules.facturacion.models.serie_correlativo_model import SerieCorrelativo

from modules.facturacion.models.producto_model import Producto
from modules.facturacion.models.personeria_model import Personeria


from .. import bp 

@bp.route('/operacion_reporte_excel')
def operacion_reporte_excel():
    ahora = datetime.now()
    periodo_actual = ahora.strftime("%m-%Y")
    periodo = request.args.get('periodo', periodo_actual)

    try:
        mes, anio = periodo.split('-')
        mes = int(mes)
        anio = int(anio)
        if not (1 <= mes <= 12) or anio < 1900:
            raise ValueError
    except (ValueError, AttributeError):
        ahora = datetime.now()
        mes = ahora.month
        anio = ahora.year

    operaciones = Operacion.get_all_by_periodo2(mes, anio)['data']

    # Crear libro
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ventas {periodo}"

    # === Título del reporte (fila 1) ===
    ws.merge_cells('A1:F1')
    ws['A1'] = f"REPORTE DE VENTAS - PERÍODO {periodo}"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="2E5984")  # Azul oscuro
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # === Encabezados (fila 2) ===
    headers = ["Serie-Número", "Fecha", "Tipo", "RUC/DNI", "Cliente", "Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")  # Azul medio
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # === Estilos de borde ===
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar borde a encabezados
    for col in range(1, 7):
        ws.cell(row=2, column=col).border = thin_border

    # === Llenar datos ===
    cont = 3
    if operaciones:
        for item in operaciones:
            # Serie-Número
            ws[f'A{cont}'] = f"{item.serie}-{str(item.correlativo).zfill(6)}"
            # Fecha
            ws[f'B{cont}'] = item.fecha_emision.strftime("%d/%m/%Y")
            # Tipo
            tipo_map = {'01': 'Factura', '03': 'Boleta', '07': 'Nota de Crédito'}
            ws[f'C{cont}'] = tipo_map.get(item.tipo_comprobante, item.tipo_comprobante)
            # RUC/DNI
            ws[f'D{cont}'] = str(item.numero_doc) if item.numero_doc else ""
            # Cliente
            ws[f'E{cont}'] = str(item.razon_social) if item.razon_social else ""
            # Total
            total = float(item.mto_imp_venta) if item.mto_imp_venta else 0.0
            ws[f'F{cont}'] = total

            # Aplicar borde a cada celda de la fila
            for col in range(1, 7):
                ws.cell(row=cont, column=col).border = thin_border
                if col == 6:  # Columna "Total"
                    ws.cell(row=cont, column=col).number_format = '#,##0.00'

            cont += 1
    else:
        # Si no hay datos
        ws.merge_cells(f'A{cont}:F{cont}')
        ws[f'A{cont}'] = "No se encontraron registros."
        ws[f'A{cont}'].alignment = Alignment(horizontal="center")
        for col in range(1, 7):
            ws.cell(row=cont, column=col).border = thin_border

    # === Ajustar ancho de columnas ===
    column_widths = {
        'A': 18,  # Serie-Número
        'B': 12,  # Fecha
        'C': 16,  # Tipo
        'D': 14,  # RUC/DNI
        'E': 30,  # Cliente
        'F': 14   # Total
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # === Guardar en buffer ===
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Ventas_{periodo}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
